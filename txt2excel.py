# !/usr/bin/python3
# coding:utf-8

import openpyxl

def deal_data(src,data,l,c):
    start=data.find(src)
 #   print("start ",start)
    if start==-1:
        return -1
    start+=len(src)
    if data[start:start+1]=='[':
        start+=1
        end=data.find("]",start)
    else:
        end=data.find(" ",start)
        print(start,end)
        if start==end:
            end=data.find(" ",start+1)
            print("空格开头",data[start:end]==" ")
        if -1==end:
            end=len(data)
    data[start:end].lstrip()
    print(src,data[start:end])
    #print(data[start:end],start,end)
    table.cell(l,c,data[start:end]) # 行，列，值 这里是从1开始计数的
   
    return 0


# 打开文件
fo = open("911tmp", "r+")
print("文件名为: ", fo.name)

data = openpyxl.Workbook() # 新建工作簿
data.create_sheet('Sheet1') # 添加页
table = data.active # 获得当前活跃的工作页，默认为第一个工作页
# 调整列宽
table.column_dimensions['A'].width = 5.0
table.column_dimensions['B'].width = 10.0
table.column_dimensions['C'].width = 15.0
table.column_dimensions['D'].width = 30.0
table.column_dimensions['E'].width = 15.0
table.column_dimensions['F'].width = 20.0
table.column_dimensions['G'].width = 40.0
table.column_dimensions['H'].width = 20.0
table.column_dimensions['I'].width = 30.0
table.column_dimensions['J'].width = 10.0
table.column_dimensions['K'].width = 10.0
table.column_dimensions['L'].width = 10.0
table.column_dimensions['N'].width = 20.0
i=1
j=1
table.cell(i,j,"序号") # 行，列，值 这里是从1开始计数的
j+=1
table.cell(i,j,"姓名") # 行，列，值 这里是从1开始计数的
j+=1
table.cell(i,j,"手机号") # 行，列，值 这里是从1开始计数的
j+=1
table.cell(i,j,"公司名称") # 行，列，值 这里是从1开始计数的
j+=1
table.cell(i,j,"职务") # 行，列，值 这里是从1开始计数的
j+=1
table.cell(i,j,"身份证号") # 行，列，值 这里是从1开始计数的
j+=1
table.cell(i,j,"收货地址") # 行，列，值 这里是从1开始计数的
j+=1
table.cell(i,j,"开票信息-税号") # 行，列，值 这里是从1开始计数的
j+=1
table.cell(i,j,"开票信息-发票抬头") # 行，列，值 这里是从1开始计数的
j+=1
table.cell(i,j,"缴费金额") # 行，列，值 这里是从1开始计数的
j+=1
table.cell(i,j,"缴费账户") # 行，列，值 这里是从1开始计数的
j+=1
table.cell(i,j,"状态") # 行，列，值 这里是从1开始计数的
j+=1
table.cell(i,j,"内部推荐人") # 行，列，值 这里是从1开始计数的
j+=1
table.cell(i,j,"备注") # 行，列，值 这里是从1开始计数的
j=1
i+=1

for line in fo.readlines():                          #依次读取每行  
    line = line.strip()  
    if len(line)<=3:
        continue                           #去掉每行头尾空白  
    print ("读取的数据为: %s" % (line))
    print("len :",len(line))

    table.cell(i,j,i-1) # 行，列，值 这里是从1开始计数的
    j+=1    

    ret=deal_data("姓名：",line,i,j)
    if ret==-1:
        print("no extis!!!")
    j+=1

    ret=deal_data("手机号：",line,i,j)
    if ret==-1:
        print("no extis!!!")
    j+=1
    
    ret=deal_data("公司名称：",line,i,j)
    if ret==-1:
        print("no extis!!!")
    j+=1
    
    ret=deal_data("职务：",line,i,j)
    if ret==-1:
        print("no extis!!!")
    j+=1
    
    ret=deal_data("身份证号（仅用于制作结业证书使用）：",line,i,j)
    if ret==-1:
        ret=deal_data("身份证号：",line,i,j)
        if ret==-1:
            ret=deal_data("身份证号",line,i,j)
            if ret==-1:
                print("no extis!!!")
    j+=1

    ret=deal_data("收货地址（仅用于快递寄送结业证书使用）：",line,i,j)
    if ret==-1: 
        ret=deal_data("收货地址：",line,i,j)   
        if ret==-1:
            print("no extis!!!")
    j+=1

    ret=deal_data("开票信息-税号：",line,i,j)
    if ret==-1:
        print("no extis!!!")
    j+=1

    ret=deal_data("开票信息-发票抬头：",line,i,j)
    if ret==-1:
        print("no extis!!!")
    j+=1
    
    ret=deal_data("缴费金额：",line,i,j)
    if ret==-1:
        print("no extis!!!")
    j+=1
    
    ret=deal_data("缴费账户：",line,i,j)
    if ret==-1:
        print("no extis!!!")
    j+=1

    ret=deal_data("状态：",line,i,j)
    if ret==-1:
        print("no extis!!!")
    j+=1

    ret=deal_data("内部推荐人：",line,i,j)
    if ret==-1:
        print("no extis!!!") 
    j+=1

    ret=deal_data("备注：",line,i,j)
    if ret==-1:
        print("no extis!!!")      

    j=1
    i+=1
    print("line:",i)

# 关闭文件
data.save('excel_test.xlsx') # 一定要保存
fo.close()


